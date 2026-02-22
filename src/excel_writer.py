"""
Excel report generation using openpyxl.
Generates a workbook with 3 sheets:
  1. Policy Summary
  2. No Withdrawal IRR Analysis
  3. With Withdrawal IRR Analysis
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

from .config import PolicyConfig


def create_excel_report(config: PolicyConfig, irr_results: list, output_path: str):
    """Generate Excel workbook with IRR analysis."""
    wb = Workbook()

    _write_summary_sheet(wb, config)
    _write_no_withdrawal_sheet(wb, config, irr_results)
    if config.withdrawal_data:
        _write_withdrawal_sheet(wb, config, irr_results)

    wb.save(output_path)


def _get_styles(config: PolicyConfig):
    """Create reusable styles based on brand config."""
    primary = config.brand.primary_color.replace('#', '')
    return {
        'header_fill': PatternFill(start_color=primary, fill_type="solid"),
        'header_font': Font(color="FFFFFF", bold=True, size=11, name="Arial"),
        'title_font': Font(color=primary, bold=True, size=16, name="Arial"),
        'subtitle_font': Font(color="333333", bold=True, size=12, name="Arial"),
        'label_font': Font(color="666666", size=11, name="Arial"),
        'value_font': Font(color="1A1A1A", bold=True, size=11, name="Arial"),
        'data_font': Font(size=10, name="Arial"),
        'irr_positive_font': Font(color="2E7D32", size=10, name="Arial", bold=True),
        'irr_negative_font': Font(color="C62828", size=10, name="Arial", bold=True),
        'irr_na_font': Font(color="9E9E9E", size=10, name="Arial", italic=True),
        'year_highlight': PatternFill(start_color="FFFFF0", fill_type="solid"),
        'age_highlight': PatternFill(start_color="F0F8FF", fill_type="solid"),
        'thin_border': Border(
            left=Side(style='thin', color='DDDDDD'),
            right=Side(style='thin', color='DDDDDD'),
            top=Side(style='thin', color='DDDDDD'),
            bottom=Side(style='thin', color='DDDDDD'),
        ),
    }


def _write_summary_sheet(wb: Workbook, config: PolicyConfig):
    """Sheet 1: Policy Summary."""
    ws = wb.active
    ws.title = "保单摘要 Summary"
    styles = _get_styles(config)
    pi = config.policy_info

    # Column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 40

    # Title
    ws.merge_cells('B2:C2')
    cell = ws['B2']
    cell.value = pi.product_name
    cell.font = styles['title_font']

    ws.merge_cells('B3:C3')
    cell = ws['B3']
    cell.value = pi.product_name_en
    cell.font = Font(color="666666", size=12, name="Arial")

    # Summary data
    summary_items = [
        ("保险公司 Insurer", pi.insurer),
        ("受保人 Insured", pi.insured_name),
        ("投保年龄 Age at Issue", str(pi.age_at_issue)),
        ("性别 Gender", "男 Male" if pi.gender == "M" else "女 Female"),
        ("保单货币 Currency", pi.currency),
        ("年缴保费 Annual Premium", f"{pi.currency_symbol}{pi.annual_premium:,.0f}"),
        ("缴费年期 Payment Period", f"{pi.payment_years} years"),
        ("总保费 Total Premium", f"{pi.currency_symbol}{pi.total_premium:,.0f}"),
        ("保障类型 Coverage", pi.coverage_type),
        ("计划日期 Plan Date", pi.plan_date),
    ]

    row = 5
    for label, value in summary_items:
        ws.cell(row=row, column=2, value=label).font = styles['label_font']
        ws.cell(row=row, column=3, value=value).font = styles['value_font']
        row += 1

    # Withdrawal info
    if config.withdrawal_data:
        row += 1
        ws.merge_cells(f'B{row}:C{row}')
        ws.cell(row=row, column=2, value="现金提取方案 Withdrawal Plan").font = styles['subtitle_font']
        row += 1

        # Find first and last withdrawal year and amounts
        first_wd = next((wd for wd in config.withdrawal_data if wd.withdrawal_amount > 0), None)
        if first_wd:
            ws.cell(row=row, column=2, value="首次提取年期 First Withdrawal").font = styles['label_font']
            ws.cell(row=row, column=3, value=f"Year {first_wd.year}").font = styles['value_font']
            row += 1
            ws.cell(row=row, column=2, value="每年提取金额 Annual Amount").font = styles['label_font']
            ws.cell(row=row, column=3, value=f"{pi.currency_symbol}{first_wd.withdrawal_amount:,.0f}").font = styles['value_font']
            row += 1

            total_withdrawals = sum(wd.withdrawal_amount for wd in config.withdrawal_data)
            ws.cell(row=row, column=2, value="累计提取金额 Total Withdrawals").font = styles['label_font']
            ws.cell(row=row, column=3, value=f"{pi.currency_symbol}{total_withdrawals:,.0f} (over {sum(1 for wd in config.withdrawal_data if wd.withdrawal_amount > 0)} years)").font = styles['value_font']


def _format_irr_cell(ws, row, col, irr_value, styles):
    """Format an IRR cell with color coding."""
    cell = ws.cell(row=row, column=col)
    if irr_value is None:
        cell.value = "N/A"
        cell.font = styles['irr_na_font']
        cell.alignment = Alignment(horizontal='center')
    else:
        cell.value = irr_value
        cell.number_format = '0.00%'
        if irr_value >= 0:
            cell.font = styles['irr_positive_font']
        else:
            cell.font = styles['irr_negative_font']
        cell.alignment = Alignment(horizontal='right')


def _write_no_withdrawal_sheet(wb: Workbook, config: PolicyConfig, irr_results: list):
    """Sheet 2: No Withdrawal IRR Analysis."""
    ws = wb.create_sheet("不提取 No Withdrawal")
    styles = _get_styles(config)
    pi = config.policy_info

    headers = [
        "Year\n年期", "Age\n年龄", f"Cumulative Premium\n累计保费 ({pi.currency})",
        f"Guaranteed CV (A)\n保证现金价值 ({pi.currency})",
        f"Rev. Bonus (B)\n复归红利 ({pi.currency})",
        f"Terminal Div (C)\n终期分红 ({pi.currency})",
        f"Total Surrender\n退保总额 ({pi.currency})",
        "IRR (Guaranteed)\n保证IRR",
        "IRR (Total)\n预期IRR",
        f"Death Benefit\n身故赔偿 ({pi.currency})",
    ]

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = styles['header_fill']
        cell.font = styles['header_font']
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = styles['thin_border']

    # Column widths
    widths = [8, 8, 18, 18, 16, 16, 18, 16, 16, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Header row height
    ws.row_dimensions[1].height = 40

    # Write data
    highlight_years = set(config.display.highlight_years)
    highlight_ages = set(config.display.highlight_ages)

    for i, (rec, irr) in enumerate(zip(config.yearly_data, irr_results)):
        row = i + 2
        is_year_highlight = rec.year in highlight_years
        is_age_highlight = rec.age in highlight_ages

        # Data cells
        data = [
            rec.year, rec.age, rec.cumulative_premium,
            rec.guaranteed_cash_value, rec.reversionary_bonus,
            rec.terminal_dividend, rec.total_surrender_value,
        ]

        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = styles['data_font']
            cell.border = styles['thin_border']
            if col >= 3:
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right')
            else:
                cell.alignment = Alignment(horizontal='center')

            # Apply highlight
            if is_year_highlight:
                cell.fill = styles['year_highlight']
            elif is_age_highlight:
                cell.fill = styles['age_highlight']

        # IRR cells
        _format_irr_cell(ws, row, 8, irr['irr_no_withdrawal_guaranteed'], styles)
        _format_irr_cell(ws, row, 9, irr['irr_no_withdrawal_total'], styles)

        # Death benefit
        cell = ws.cell(row=row, column=10, value=rec.total_death_benefit)
        cell.font = styles['data_font']
        cell.number_format = '#,##0'
        cell.alignment = Alignment(horizontal='right')
        cell.border = styles['thin_border']

        # Apply highlight to IRR and death benefit cells too
        for col in [8, 9, 10]:
            if is_year_highlight:
                ws.cell(row=row, column=col).fill = styles['year_highlight']
            elif is_age_highlight:
                ws.cell(row=row, column=col).fill = styles['age_highlight']

    # Freeze panes
    ws.freeze_panes = 'A2'

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(config.yearly_data) + 1}"


def _write_withdrawal_sheet(wb: Workbook, config: PolicyConfig, irr_results: list):
    """Sheet 3: With Withdrawal IRR Analysis."""
    ws = wb.create_sheet("提取 With Withdrawal")
    styles = _get_styles(config)
    pi = config.policy_info

    headers = [
        "Year\n年期", "Age\n年龄",
        f"Withdrawal\n当年提取 ({pi.currency})",
        f"Cumulative Withdrawals\n累计提取 ({pi.currency})",
        f"Remaining Guaranteed (A)\n剩余保证 ({pi.currency})",
        f"Remaining Total\n剩余退保总额 ({pi.currency})",
        "IRR (Guaranteed)\n保证IRR",
        "IRR (Total)\n预期IRR",
    ]

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = styles['header_fill']
        cell.font = styles['header_font']
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = styles['thin_border']

    # Column widths
    widths = [8, 8, 16, 20, 20, 20, 16, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 40

    # Build withdrawal lookup and calculate cumulative
    wd_by_year = {wd.year: wd for wd in config.withdrawal_data}
    cumulative_wd = 0.0

    highlight_years = set(config.display.highlight_years)
    highlight_ages = set(config.display.highlight_ages)

    for i, irr in enumerate(irr_results):
        year = irr['year']
        age = irr['age']

        if year not in wd_by_year:
            continue

        wd_rec = wd_by_year[year]
        cumulative_wd += wd_rec.withdrawal_amount

        row = ws.max_row + 1 if ws.max_row > 1 else 2
        is_year_highlight = year in highlight_years
        is_age_highlight = age in highlight_ages

        data = [
            year, age,
            wd_rec.withdrawal_amount,
            cumulative_wd,
            wd_rec.remaining_surrender_guaranteed,
            wd_rec.remaining_surrender_total,
        ]

        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = styles['data_font']
            cell.border = styles['thin_border']
            if col >= 3:
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right')
            else:
                cell.alignment = Alignment(horizontal='center')

            if is_year_highlight:
                cell.fill = styles['year_highlight']
            elif is_age_highlight:
                cell.fill = styles['age_highlight']

        # IRR cells
        _format_irr_cell(ws, row, 7, irr['irr_withdrawal_guaranteed'], styles)
        _format_irr_cell(ws, row, 8, irr['irr_withdrawal_total'], styles)

        for col in [7, 8]:
            if is_year_highlight:
                ws.cell(row=row, column=col).fill = styles['year_highlight']
            elif is_age_highlight:
                ws.cell(row=row, column=col).fill = styles['age_highlight']

    # Freeze panes
    ws.freeze_panes = 'A2'

    # Auto-filter
    max_row = ws.max_row
    if max_row > 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max_row}"
